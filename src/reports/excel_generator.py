from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADER_FILL = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
HEADER_FONT = Font(color="ffffff", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def export_to_excel(data: list[dict], headers: list[str], output_path: str | Path, sheet_name: str = "Report"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_row = []
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        header_row.append(cell)

    for row_idx, record in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            key = header.lower().replace(" ", "_").replace("/", "_")
            value = record.get(key, record.get(header, ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if isinstance(value, (int, float)) else "left")

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for row_idx in range(1, len(data) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[chr(64 + col_idx)].width = min(max_len + 4, 40)

    wb.save(str(output_path))
    return output_path


def export_table_widget_to_excel(table_widget, output_path: str | Path):
    headers = []
    for col in range(table_widget.columnCount()):
        header_item = table_widget.horizontalHeaderItem(col)
        headers.append(header_item.text() if header_item else f"Col{col}")

    data = []
    for row in range(table_widget.rowCount()):
        record = {}
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            record[headers[col]] = item.text() if item else ""
        data.append(record)

    return export_to_excel(data, headers, output_path)
